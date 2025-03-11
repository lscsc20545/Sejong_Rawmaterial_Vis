# -*- coding: utf-8 -*-

import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import numpy as np
import warnings
import platform
import plotly.express as px
from scipy import stats
import streamlit.components.v1 as components


# openpyxl 사용
import openpyxl
try:
    # openpyxl로 시도
    import pandas as pd
    pd.read_excel('data/sample_data.xlsx', engine='openpyxl')
except:
    # 실패하면 pandas로 폴백
    import pandas as pd
    pd.read_excel('data/sample_data.xlsx')

from datetime import datetime, timedelta

def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    # 스타일 추가
    st.markdown(
        """
        <style>
        div[data-testid="stTextInput"] {
            padding-top: 2rem;
            margin-top: 1rem;
        }
        </style>
        """, 
        unsafe_allow_html=True
    )
    
    if "password_correct" not in st.session_state:
        st.text_input(
            "비밀번호 4자리를 입력하세요", 
            type="password", 
            on_change=password_entered, 
            key="password",
            label_visibility="visible"
        )
        return False
    elif not st.session_state["password_correct"]:
        st.text_input(
            "비밀번호를 입력하세요", 
            type="password", 
            on_change=password_entered, 
            key="password",
            label_visibility="visible"
        )
        st.error("😕 비밀번호가 올바르지 않습니다.")
        return False
    else:
        return True


warnings.filterwarnings('ignore')

# 페이지 설정
st.set_page_config(page_title="장섬유 조성 관리 대시보드", layout="wide")

# 사이드바 너비 조정 및 스타일 개선
st.markdown(
    """
    <style>
    [data-testid="stSidebar"][aria-expanded="true"]{
        min-width: 300px;
        max-width: 300px;
    }
    .stDataFrame {
        width: 100% !important;
    }
    .dataframe {
        width: 100% !important;
    }
    /* 위젯 간격 줄이기 */
    .block-container {
        padding-top: 3rem;  /* 상단 여백 늘림 */
        padding-bottom: 1rem;
    }
    div.row-widget.stRadio > div {
        flex-direction: row;
        align-items: center;
    }
    div.row-widget.stRadio > div > label {
        margin: 0 0.5rem;
    }
    /* 타이틀 스타일 개선 */
    h1, h2, h3, h4, h5, h6 {
        padding-top: 1.5rem;
        margin-top: 0.8rem;
    }
    /* 라디오 버튼 스타일 개선 */
    div[data-testid="stRadio"] {
        padding-top: 1rem;
        margin-top: 0.5rem;
    }
    /* 위젯 카드 스타일 */
    div.css-1r6slb0.e1tzin5v2 {
        background-color: #f5f5f5;
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin-bottom: 10px;
    }
    /* 메트릭 카드 스타일 */
    .stMetric {
        background-color: white;
        border-radius: 8px;
        padding: 15px !important;
        margin-bottom: 0.8rem !important;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        border-left: 4px solid #4e8cff;
    }
    /* 메트릭 레이블 스타일 */
    .stMetric > div:first-child {
        color: #555;
        font-weight: 600;
    }
    /* 메트릭 값 스타일 */
    .stMetric > div:nth-child(2) {
        font-size: 1.3rem;
        font-weight: bold;
        color: #1f77b4;
    }
    /* 이상치 정보 스타일 */
    .anomaly-box {
        background-color: #f8f9fa;
        border-radius: 10px;
        padding: 15px;
        margin-top: 20px;
        border-left: 4px solid #ff7043;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
    }
    /* 부적합 정보 스타일 */
    .incompatible-box {
        background-color: #f8f9fa;
        border-radius: 10px;
        padding: 15px;
        margin-top: 10px;
        border-left: 4px solid #42a5f5;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
    }
    </style>
    """, 
    unsafe_allow_html=True
)


def initialize_session_state():
    """
    세션 상태 초기화 함수
    """
    if 'selected_item' not in st.session_state:
        st.session_state.selected_item = None
    
    if 'selected_data' not in st.session_state:
        st.session_state.selected_data = None
        
    # 라디오 버튼 선택 상태를 위한 세션 변수
    if 'tab_selection' not in st.session_state:
        st.session_state.tab_selection = "전체 현황"

# 데이터 로드 함수
@st.cache_data
def load_sample_data():
    """샘플 데이터 로드 (openpyxl 사용)"""
    try:
        # 샘플 데이터 파일 경로
        sample_file = "data/sample_data.xlsx"
        all_data = {}
        
        # openpyxl로 엑셀 파일 열기
        import openpyxl
        wb = openpyxl.load_workbook(sample_file)
        
        # 모든 시트 처리
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 데이터 범위 읽기
            values = list(sheet.values)
            
            # 헤더가 있다고 가정하고 데이터프레임 생성
            if values and len(values) > 1:
                headers = values[0]
                data = values[1:]
                df = pd.DataFrame(data, columns=headers)
                
                # 인덱스 컬럼이 없으면 추가
                if '날짜' not in df.columns:
                    df = df.reset_index()
                    df = df.rename(columns={'index': '날짜'})
                
                # 숫자 데이터 변환
                numeric_columns = ['실측', '배합', '상한선', '하한선']
                for col in numeric_columns:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')
                
                # 날짜 변환
                if pd.api.types.is_numeric_dtype(df['날짜']):
                    df['날짜'] = pd.TimedeltaIndex(df['날짜'], unit='D') + pd.Timestamp('1899-12-30')
                else:
                    df['날짜'] = pd.to_datetime(df['날짜'], errors='coerce')
                
                # 결측치 처리
                df = df.dropna(subset=['날짜'])
                
                # 시트 이름을 구분으로 추가
                df['sheet_name'] = sheet_name
                
                all_data[sheet_name] = df
        
        return all_data
    except Exception as e:
        st.error(f"샘플 데이터 로드 중 오류 발생: {str(e)}")
        return None



def load_uploaded_data(uploaded_file):
    """업로드된 파일 데이터 로드 (openpyxl 사용)"""
    try:
        # 임시 파일로 저장
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded_file.getvalue())
            temp_path = tmp.name
        
        all_data = {}
        
        # openpyxl로 엑셀 파일 열기
        import openpyxl
        wb = openpyxl.load_workbook(temp_path)
        
        # 모든 시트 처리
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 데이터 범위 읽기
            values = list(sheet.values)
            
            # 헤더가 있다고 가정하고 데이터프레임 생성
            if values and len(values) > 1:
                headers = values[0]
                data = values[1:]
                df = pd.DataFrame(data, columns=headers)
                
                # 인덱스 컬럼이 없으면 추가
                if '날짜' not in df.columns:
                    df = df.reset_index()
                    df = df.rename(columns={'index': '날짜'})
                
                # 숫자 데이터 변환
                numeric_columns = ['실측', '배합', '상한선', '하한선']
                for col in numeric_columns:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')
                
                # 날짜 변환
                if pd.api.types.is_numeric_dtype(df['날짜']):
                    df['날짜'] = pd.TimedeltaIndex(df['날짜'], unit='D') + pd.Timestamp('1899-12-30')
                else:
                    df['날짜'] = pd.to_datetime(df['날짜'], errors='coerce')
                
                # 결측치 처리
                df = df.dropna(subset=['날짜'])
                
                # 시트 이름을 구분으로 추가
                df['sheet_name'] = sheet_name
                
                all_data[sheet_name] = df
        
        # 임시 파일 삭제
        import os
        os.unlink(temp_path)
        
        return all_data
    except Exception as e:
        st.error(f"파일 업로드 중 오류 발생: {str(e)}")
        return None




def calculate_process_capability(data, ucl, lcl, sigma_level=3):
    mean = data.mean()
    std = data.std()
    
    # 공정능력지수 계산
    cp = (ucl - lcl) / (6 * std) if std != 0 else float('inf')
    cpu = (ucl - mean) / (3 * std) if std != 0 else float('inf')
    cpl = (mean - lcl) / (3 * std) if std != 0 else float('inf')
    cpk = min(cpu, cpl)
    
    # 예상불량률 계산 (ppm 단위)
    z_upper = (ucl - mean) / std if std != 0 else float('inf')
    z_lower = (mean - lcl) / std if std != 0 else float('inf')
    ppm_upper = stats.norm.sf(z_upper) * 1000000
    ppm_lower = stats.norm.sf(z_lower) * 1000000
    total_ppm = ppm_upper + ppm_lower
    
    return {
        'Cp': cp,
        'Cpu': cpu,
        'Cpl': cpl,
        'Cpk': cpk,
        'PPM': total_ppm
    }

def main():
    # 비밀번호 체크
    if not check_password():
        # 로그인 화면에서도 타이틀 표시
        st.title("장섬유 조성 관리 대시보드")
        return
    
    # 로그인 성공 후 메인 화면에도 타이틀 표시
    st.title("장섬유 조성 관리 대시보드")
    
    # 사이드바에 파일 업로드 기능 추가
    with st.sidebar:
        st.markdown("### 데이터 업로드")
        uploaded_file = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=['xlsx'])
        
        # 초기화 버튼
        if st.button("샘플 데이터로 초기화"):
            st.session_state.data = load_sample_data()
            st.rerun()
    
# 데이터 로드 로직
    if 'data' not in st.session_state:
        st.session_state.data = load_sample_data()
    
    # 파일이 업로드되면 해당 데이터 사용
    if uploaded_file is not None:
        uploaded_data = load_uploaded_data(uploaded_file)
        if uploaded_data is not None:
            st.session_state.data = uploaded_data
    
    all_data = st.session_state.data  # load_data() 대신 세션 상태에서 데이터 가져오기
    
    if not all_data:
        st.error("데이터를 로드할 수 없습니다.")
        return
    
    # 사이드바 설정
    st.sidebar.header("필터 설정")
    
    # 시트(제품) 선택
    sheet_names = list(all_data.keys())
    selected_sheet = st.sidebar.selectbox("제품 선택", sheet_names)
    
    # 선택된 시트의 데이터 가져오기
    df = all_data[selected_sheet]
    
    # 데이터 정렬 및 최근 30개 데이터 기본 선택
    df = df.sort_values('날짜', ascending=False)
    
    # 데이터 표시 개수 선택
    display_option = st.sidebar.radio(
        "데이터 표시 범위",
        ["최근 30개", "최근 90개", "모든 데이터", "날짜 범위 지정"]
    )
    
    # 선택한 옵션에 따라 데이터 필터링
    if display_option == "최근 30개":
        # 항목별로 최근 30개 데이터 선택
        filtered_items = []
        for item in df['항목'].unique():
            item_data = df[df['항목'] == item].sort_values('날짜', ascending=False).head(30)
            filtered_items.append(item_data)
        
        filtered_df = pd.concat(filtered_items)
        date_min = filtered_df['날짜'].min()
        date_max = filtered_df['날짜'].max()
        
    elif display_option == "최근 90개":
        # 항목별로 최근 90개 데이터 선택
        filtered_items = []
        for item in df['항목'].unique():
            item_data = df[df['항목'] == item].sort_values('날짜', ascending=False).head(90)
            filtered_items.append(item_data)
        
        filtered_df = pd.concat(filtered_items)
        date_min = filtered_df['날짜'].min()
        date_max = filtered_df['날짜'].max()
        
    elif display_option == "모든 데이터":
        filtered_df = df
        date_min = filtered_df['날짜'].min()
        date_max = filtered_df['날짜'].max()
        
    else:  # "날짜 범위 지정"
        # 날짜 범위 선택
        date_min = df['날짜'].min().to_pydatetime().date()
        date_max = df['날짜'].max().to_pydatetime().date()
        
        date_range = st.sidebar.date_input(
            "날짜 범위 선택",
            value=(date_min, date_max),
            min_value=date_min,
            max_value=date_max
        )
        
        # 날짜 범위가 올바르게 선택되었는지 확인
        if len(date_range) == 2:
            start_date, end_date = date_range
            filtered_df = df[
                (df['날짜'].dt.date >= start_date) &
                (df['날짜'].dt.date <= end_date)
            ]
        else:
            st.warning("날짜 범위를 올바르게 선택해주세요.")
            filtered_df = df  # 기본값으로 모든 데이터 사용
    
    # 이상치 기준 시그마 선택
    sigma = st.sidebar.slider("이상치 기준 (σ)", 1.0, 4.0, 3.0, 0.1)
    # 세션 상태 초기화
    initialize_session_state()

    # 탭 대신 라디오 버튼으로 화면 전환
    st.markdown('<div style="padding-top: 1rem;"></div>', unsafe_allow_html=True)
    tab_selection = st.radio("화면 선택", ["전체 현황", "상세 분석"], 
                            index=0 if st.session_state.tab_selection == "전체 현황" else 1,
                            horizontal=True)

    # 라디오 버튼 선택 값을 세션 상태에 저장
    st.session_state.tab_selection = tab_selection

    if tab_selection == "전체 현황":
        display_overview(all_data, filtered_df, selected_sheet, sigma)
    else:
        display_detailed_analysis(filtered_df, selected_sheet, sigma)



def display_detailed_analysis(filtered_df, selected_sheet, sigma):
    # 조성 항목 선택
    default_selection = []
    if 'selected_item' in st.session_state and st.session_state.selected_item:
        # 세션 상태에 저장된 항목이 있고 그것이 현재 데이터에 있으면 기본값으로 설정
        if st.session_state.selected_item in filtered_df['항목'].unique():
            default_selection = [st.session_state.selected_item]
    
    # 기본값이 없으면 첫 번째 항목 선택
    if not default_selection and len(filtered_df['항목'].unique()) > 0:
        default_selection = [filtered_df['항목'].unique()[0]]
    
    composition_types = st.sidebar.multiselect(
        "조성 항목 선택",
        options=filtered_df['항목'].unique(),
        default=default_selection
    )
    
    # 선택 후 세션 상태 초기화 (다음 탭 전환에 영향을 주지 않도록)
    if st.session_state.selected_item:
        st.session_state.selected_item = None
    
    if not composition_types:
        st.warning("조성 항목을 선택해주세요.")
        return

    
    # 메인 컨테이너 - 위젯 부분
    main_container = st.container()
    
    with main_container:
        if len(composition_types) == 1:
            # 단일 항목 선택 시
            st.subheader(f"{selected_sheet} 조성 추이 - {composition_types[0]}")
            item_data = filtered_df[filtered_df['항목'] == composition_types[0]]
            
            # 데이터 포인트 인덱스 생성
            item_data = item_data.sort_values('날짜')
            
            # 통계적 관리한계 계산
            mean = item_data['실측'].mean()
            std = item_data['실측'].std()
            statistical_ucl = mean + sigma * std
            statistical_lcl = mean - sigma * std
            
            # 정상 및 이상치 포인트 구분
            normal_points = item_data[abs(item_data['실측'] - mean) <= sigma * std]
            outlier_points = item_data[abs(item_data['실측'] - mean) > sigma * std]
            
            # 부적합 수 계산 (설정 상/하한선을 벗어나는 경우)
            out_of_spec = item_data[
                (item_data['실측'] > item_data['상한선']) | 
                (item_data['실측'] < item_data['하한선'])
            ]
            out_of_spec_count = len(out_of_spec)
            out_of_spec_ratio = out_of_spec_count / len(item_data) * 100 if len(item_data) > 0 else 0
            
            # 배합과 실측의 차이 계산 추가
            item_data['편차'] = item_data['실측'] - item_data['배합']
            mean_diff = item_data['편차'].mean()
            std_diff = item_data['편차'].std()
            abs_mean_diff = item_data['편차'].abs().mean()
            
            # 주요 통계 위젯 표시 - 카드 형식으로 개선
            st.markdown('<div class="card-container">', unsafe_allow_html=True)
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric(label="평균", value=f"{mean:.3f}")
            with col2:
                st.metric(label="표준편차", value=f"{std:.3f}")
            with col3:
                st.metric(label="이상치 수", value=f"{len(outlier_points)} ({len(outlier_points)/len(item_data)*100:.1f}%)")
            with col4:
                st.metric(label="부적합 수 (비율)", value=f"{out_of_spec_count} ({out_of_spec_ratio:.1f}%)")
            with col5:
                st.metric(label="평균 편차", value=f"{mean_diff:.3f}")
            st.markdown('</div>', unsafe_allow_html=True)
            
        else:
            # 다중 항목 선택 시
            st.subheader(f"{selected_sheet} 조성 추이 - 다중 항목")
            
            # 통계 요약 표시 - 카드 형식으로 개선
            items_stats = []
            for item in composition_types:
                item_data = filtered_df[filtered_df['항목'] == item]
                item_data = item_data.sort_values('날짜')
                
                # 통계치 계산
                mean = item_data['실측'].mean()
                std = item_data['실측'].std()
                
                # 이상치 수 계산
                outlier_count = len(item_data[abs(item_data['실측'] - mean) > sigma * std])
                
                # 부적합 수 계산 (설정 상/하한선을 벗어나는 경우)
                out_of_spec = item_data[
                    (item_data['실측'] > item_data['상한선']) | 
                    (item_data['실측'] < item_data['하한선'])
                ]
                out_of_spec_count = len(out_of_spec)
                
                # 배합과 실측의 차이 계산 추가
                item_data['편차'] = item_data['실측'] - item_data['배합']
                mean_diff = item_data['편차'].mean()
                std_diff = item_data['편차'].std()
                abs_mean_diff = item_data['편차'].abs().mean()
                
                # 통계 정보 저장
                items_stats.append({
                    '항목': item,
                    '평균': mean,
                    '표준편차': std,
                    '이상치 수': outlier_count,
                    '이상치 비율(%)': outlier_count/len(item_data)*100 if len(item_data) > 0 else 0,
                    '부적합 수': out_of_spec_count,
                    '부적합 비율(%)': out_of_spec_count/len(item_data)*100 if len(item_data) > 0 else 0,
                    '평균 편차': mean_diff,
                    '절대 평균 편차': abs_mean_diff
                })
            
            # 색상 팔레트 설정
            colors = px.colors.qualitative.Plotly
            
            # 항목별 카드 컨테이너 시작
            st.markdown('<div class="stats-container" style="margin-top: 20px;">', unsafe_allow_html=True)
            stats_cols = st.columns(len(composition_types))
            for i, item in enumerate(composition_types):
                with stats_cols[i]:
                    item_stats = [s for s in items_stats if s['항목'] == item][0]
                    # 카드 스타일 적용
                    st.markdown(f"""
                    <div style="background-color: #f8f9fa; border-radius: 10px; padding: 15px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); border-left: 5px solid {colors[i % len(colors)]};">
                        <h4 style="margin-top: 0;">{item}</h4>
                    </div>
                    """, unsafe_allow_html=True)
                    st.metric(label=f"평균", value=f"{item_stats['평균']:.3f}")
                    st.metric(label="이상치", value=f"{int(item_stats['이상치 수'])} ({item_stats['이상치 비율(%)']:.1f}%)")
                    st.metric(label="부적합 수", value=f"{int(item_stats['부적합 수'])} ({item_stats['부적합 비율(%)']:.1f}%)")
                    st.metric(label="평균 편차", value=f"{item_stats['평균 편차']:.3f}")
            st.markdown('</div>', unsafe_allow_html=True)
    
    # 차트 컨테이너 - 위젯 아래에 위치하도록 변경
    chart_container = st.container()
    
    with chart_container:
        st.markdown(f"""
        <div style="background-color: #f0f2f6; border-radius: 10px; padding: 15px; margin-top: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
            <h3 style="margin-top: 0;">조성 추이 그래프</h3>
        </div>
        """, unsafe_allow_html=True)
        
        if len(composition_types) == 1:
            item_data = filtered_df[filtered_df['항목'] == composition_types[0]]
            item_data = item_data.sort_values('날짜')
            
            # 통계적 관리한계 계산
            mean = item_data['실측'].mean()
            std = item_data['실측'].std()
            statistical_ucl = mean + sigma * std
            statistical_lcl = mean - sigma * std
            
            # 정상 및 이상치 포인트 구분
            outlier_points = item_data[abs(item_data['실측'] - mean) > sigma * std]
            
            fig = go.Figure()
            
            # 날짜를 등간격으로 표시하기 위한 처리
            # 날짜를 정렬하고 인덱스 부여
            item_data = item_data.sort_values('날짜').reset_index(drop=True)
            
            # 실측값 선과 정상 포인트를 하나의 트레이스로 통합
            fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                y=item_data['실측'],
                name='실측',
                mode='lines+markers',
                line=dict(color='rgb(0, 0, 255)', width=2),
                marker=dict(
                    color='rgb(0, 0, 255)',
                    size=8
                ),
                text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                hovertemplate='날짜: %{text}<br>실측: %{y:.3f}<extra></extra>',
                customdata=item_data.index  # 인덱스 정보 저장
            ))
            
            # 이상치 데이터 포인트
            if not outlier_points.empty:
                # 이상치 데이터의 인덱스 찾기
                outlier_indices = []
                outlier_values = []
                outlier_dates = []
                outlier_ids = []
                
                for idx, row in item_data.iterrows():
                    if abs(row['실측'] - mean) > sigma * std:
                        outlier_indices.append(idx)
                        outlier_values.append(row['실측'])
                        outlier_dates.append(row['날짜'].strftime('%Y-%m-%d'))
                        outlier_ids.append(idx)  # 인덱스 저장
                
                fig.add_trace(go.Scatter(
                    x=outlier_indices,
                    y=outlier_values,
                    name='이상치',
                    mode='markers',
                    marker=dict(
                        color='rgba(255, 0, 0, 0.7)',
                        size=15,
                        line=dict(
                            color='red',
                            width=2
                        )
                    ),
                    text=outlier_dates,
                    hovertemplate='날짜: %{text}<br>실측: %{y:.3f}<extra></extra>',
                    customdata=outlier_ids  # 인덱스 정보 저장
                ))
            
            # 배합값 트레이스
            fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                y=item_data['배합'],
                name='배합',
                mode='lines',
                line=dict(dash='dash', color='rgb(255, 165, 0)', width=2),
                text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                hovertemplate='날짜: %{text}<br>배합: %{y:.3f}<extra></extra>'
            ))
            
            # 설정된 관리한계 트레이스
            if '상한선' in item_data.columns:
                fig.add_trace(go.Scatter(
                    x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                    y=item_data['상한선'],
                    name='설정 상한선',
                    mode='lines',
                    line=dict(dash='dot', color='darkgreen', width=2),
                    text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                    hovertemplate='날짜: %{text}<br>상한선: %{y:.3f}<extra></extra>'
                ))
            
            if '하한선' in item_data.columns:
                fig.add_trace(go.Scatter(
                    x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                    y=item_data['하한선'],
                    name='설정 하한선',
                    mode='lines',
                    line=dict(dash='dot', color='darkgreen', width=2),
                    text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                    hovertemplate='날짜: %{text}<br>하한선: %{y:.3f}<extra></extra>'
                ))
            
            # 통계적 관리한계 트레이스
            fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                y=[statistical_ucl] * len(item_data),
                name=f'통계적 상한선 ({sigma}σ)',
                mode='lines',
                line=dict(dash='dot', color='red', width=1.5),
                hovertemplate=f'통계적 상한선: {statistical_ucl:.3f}<extra></extra>'
            ))
            
            fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                y=[statistical_lcl] * len(item_data),
                name=f'통계적 하한선 ({sigma}σ)',
                mode='lines',
                line=dict(dash='dot', color='red', width=1.5),
                hovertemplate=f'통계적 하한선: {statistical_lcl:.3f}<extra></extra>'
            ))
            
            # X축 레이블 설정 (날짜 표시)
            # 날짜 레이블을 적절한 간격으로 표시
            n_points = len(item_data)
            
            # 표시할 날짜 레이블 수 결정 (최대 15개)
            n_labels = min(15, n_points)
            
            # 레이블 위치 계산
            if n_points <= n_labels:
                # 데이터 포인트가 적으면 모든 날짜 표시
                label_indices = list(range(n_points))
            else:
                # 데이터 포인트가 많으면 균등하게 분포
                step = n_points / n_labels
                label_indices = [int(i * step) for i in range(n_labels)]
                # 마지막 인덱스 추가
                if label_indices[-1] != n_points - 1:
                    label_indices.append(n_points - 1)
            
            # 레이블 생성
            tickvals = label_indices
            ticktext = [item_data.iloc[i]['날짜'].strftime('%y-%m-%d') for i in label_indices]
            
            # X축 설정 - 등간격으로 표시하고 날짜 레이블 추가
            fig.update_layout(
                xaxis=dict(
                    title='측정 순서',
                    tickmode='array',
                    tickvals=tickvals,
                    ticktext=ticktext,
                    tickangle=45
                ),
                yaxis_title='성분(%)',
                hovermode='x unified',
                showlegend=True,
                legend=dict(
                    yanchor="top",
                    y=0.99,
                    xanchor="left",
                    x=0.01,
                    traceorder='grouped'
                ),
                margin=dict(l=40, r=40, t=40, b=80),
                height=500
            )
            
        else:
            # 다중 항목 선택 시
            # 색상 팔레트 설정
            colors = px.colors.qualitative.Plotly
            
            fig = go.Figure()
            
            # 각 항목별로 그래프 추가
            for i, item in enumerate(composition_types):
                color = colors[i % len(colors)]
                item_data = filtered_df[filtered_df['항목'] == item].sort_values('날짜')
                
                # 등간격 X축을 위한 인덱스 생성
                item_data = item_data.reset_index(drop=True)
                
                # 통계치 계산
                mean = item_data['실측'].mean()
                std = item_data['실측'].std()
                statistical_ucl = mean + sigma * std
                statistical_lcl = mean - sigma * std
                
                # 실측값 선과 포인트를 하나의 트레이스로 통합
                fig.add_trace(go.Scatter(
                    x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                    y=item_data['실측'],
                    name=f'{item} (실측)',
                    mode='lines+markers',
                    line=dict(color=color, width=2),
                    marker=dict(
                        color=color,
                        size=8
                    ),
                    text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                    hovertemplate='항목: %{fullData.name}<br>날짜: %{text}<br>실측: %{y:.3f}<extra></extra>'
                ))
                
                # 배합값 트레이스
                fig.add_trace(go.Scatter(
                    x=list(range(len(item_data))),  # 등간격 X축을 위해 인덱스 사용
                    y=item_data['배합'],
                    name=f'{item} (배합)',
                    mode='lines',
                    line=dict(dash='dash', color=color, width=1.5),
                    opacity=0.7,
                    text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                    hovertemplate='항목: %{fullData.name}<br>날짜: %{text}<br>배합: %{y:.3f}<extra></extra>'
                ))
                
                # 각 항목별로 날짜 레이블 생성
                n_points = len(item_data)
                n_labels = min(10, n_points)  # 최대 10개 레이블
                
                if n_points <= n_labels:
                    label_indices = list(range(n_points))
                else:
                    step = n_points / n_labels
                    label_indices = [int(i * step) for i in range(n_labels)]
                    if label_indices[-1] != n_points - 1:
                        label_indices.append(n_points - 1)
                
                # 첫 번째 항목의 레이블만 사용
                if i == 0:
                    tickvals = label_indices
                    ticktext = [item_data.iloc[i]['날짜'].strftime('%y-%m-%d') for i in label_indices]
            
            # X축 설정 - 등간격으로 표시하고 날짜 레이블 추가 (첫 번째 항목 기준)
            fig.update_layout(
                xaxis=dict(
                    title='측정 순서',
                    tickmode='array',
                    tickvals=tickvals,
                    ticktext=ticktext,
                    tickangle=45
                ),
                yaxis_title='성분(%)',
                hovermode='x unified',
                showlegend=True,
                legend=dict(
                    orientation='h',
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1,
                    traceorder='grouped'
                ),
                margin=dict(l=40, r=40, t=40, b=80),
                height=600
            )
        
        # 그래프 표시 및 클릭 이벤트 설정
        selected_points = plotly_chart_with_click_event(fig, key=f"chart_{selected_sheet}")
        
        # 클릭된 포인트가 있으면 해당 데이터 표시
        if selected_points:
            point_index = selected_points[0]['pointIndex']
            curve_index = selected_points[0]['curveNumber']
            
            # 클릭된 포인트의 데이터 찾기
            if curve_index == 0:  # 실측 데이터
                selected_data = item_data.iloc[point_index]
                st.session_state.selected_data = selected_data
                st.session_state.selected_item = composition_types[0]
                
                # 상세 정보 표시
                st.markdown(f"""
                <div style="background-color: #e8f4f8; border-radius: 10px; padding: 15px; margin-top: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
                    <h4 style="margin-top: 0;">선택된 데이터 상세 정보</h4>
                    <p>날짜: {selected_data['날짜'].strftime('%Y년 %m월 %d일')}</p>
                    <p>항목: {composition_types[0]}</p>
                    <p>실측값: {selected_data['실측']:.3f}</p>
                    <p>배합값: {selected_data['배합']:.3f}</p>
                    <p>편차: {selected_data['실측'] - selected_data['배합']:.3f}</p>
                </div>
                """, unsafe_allow_html=True)



        # 배합과 실측 차이 그래프 추가
        st.markdown(f"""
        <div style="background-color: #f0f2f6; border-radius: 10px; padding: 15px; margin-top: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
            <h3 style="margin-top: 0;">배합-실측 편차 그래프</h3>
        </div>
        """, unsafe_allow_html=True)

        if len(composition_types) == 1:
            # 단일 항목 선택 시 편차 그래프
            # 편차 그래프 생성 부분 수정 (약 1079번째 줄 부근)
            # 단일 항목 선택 시 편차 그래프
            item_data = filtered_df[filtered_df['항목'] == composition_types[0]]
            item_data = item_data.sort_values('날짜').reset_index(drop=True)

            # 편차 계산
            item_data['편차'] = item_data['실측'] - item_data['배합']

            # 편차의 통계적 관리한계 계산
            diff_mean = item_data['편차'].mean()
            diff_std = item_data['편차'].std()
            diff_statistical_ucl = diff_mean + sigma * diff_std
            diff_statistical_lcl = diff_mean - sigma * diff_std

            # 편차 그래프 생성
            diff_fig = go.Figure()

            # 편차 선 그래프
            diff_fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),
                y=item_data['편차'],
                name='편차 (실측-배합)',
                mode='lines+markers',
                line=dict(color='purple', width=2),
                marker=dict(
                    color='purple',
                    size=8
                ),
                text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                hovertemplate='날짜: %{text}<br>편차: %{y:.3f}<extra></extra>'
            ))

            # 0선 추가 (기준선)
            diff_fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),
                y=[0] * len(item_data),
                name='기준선',
                mode='lines',
                line=dict(color='black', width=1, dash='dash')
            ))

            # 평균 편차선 추가
            diff_fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),
                y=[diff_mean] * len(item_data),
                name=f'평균 편차: {diff_mean:.3f}',
                mode='lines',
                line=dict(color='red', width=1.5)
            ))

            # 통계적 상한선, 하한선 추가
            diff_fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),
                y=[diff_statistical_ucl] * len(item_data),
                name=f'통계적 상한선 ({sigma}σ)',
                mode='lines',
                line=dict(color='red', width=1.5, dash='dot')
            ))

            diff_fig.add_trace(go.Scatter(
                x=list(range(len(item_data))),
                y=[diff_statistical_lcl] * len(item_data),
                name=f'통계적 하한선 ({sigma}σ)',
                mode='lines',
                line=dict(color='red', width=1.5, dash='dot')
            ))

            
            # 날짜 레이블 설정
            n_points = len(item_data)
            n_labels = min(15, n_points)
            
            if n_points <= n_labels:
                label_indices = list(range(n_points))
            else:
                step = n_points / n_labels
                label_indices = [int(i * step) for i in range(n_labels)]
                if label_indices[-1] != n_points - 1:
                    label_indices.append(n_points - 1)
            
            tickvals = label_indices
            ticktext = [item_data.iloc[i]['날짜'].strftime('%y-%m-%d') for i in label_indices]
            
            # 레이아웃 설정
            diff_fig.update_layout(
                xaxis=dict(
                    title='측정 순서',
                    tickmode='array',
                    tickvals=tickvals,
                    ticktext=ticktext,
                    tickangle=45
                ),
                yaxis_title='편차 (실측-배합)',
                hovermode='x unified',
                showlegend=True,
                legend=dict(
                    yanchor="top",
                    y=0.99,
                    xanchor="left",
                    x=0.01
                ),
                margin=dict(l=40, r=40, t=40, b=80),
                height=400
            )
            
        else:
            # 다중 항목 선택 시 편차 그래프
            diff_fig = go.Figure()
            
            # 색상 팔레트 설정
            colors = px.colors.qualitative.Plotly
            
            # 각 항목별로 편차 그래프 추가
            for i, item in enumerate(composition_types):
                color = colors[i % len(colors)]
                item_data = filtered_df[filtered_df['항목'] == item].sort_values('날짜').reset_index(drop=True)
                
                # 편차 계산
                item_data['편차'] = item_data['실측'] - item_data['배합']
                
                # 편차 선 그래프
                diff_fig.add_trace(go.Scatter(
                    x=list(range(len(item_data))),
                    y=item_data['편차'],
                    name=f'{item} 편차',
                    mode='lines+markers',
                    line=dict(color=color, width=2),
                    marker=dict(
                        color=color,
                        size=8
                    ),
                    text=item_data['날짜'].dt.strftime('%Y-%m-%d'),
                    hovertemplate='항목: %{fullData.name}<br>날짜: %{text}<br>편차: %{y:.3f}<extra></extra>'
                ))
                
                # 평균 편차선 추가
                mean_diff = item_data['편차'].mean()
                diff_fig.add_trace(go.Scatter(
                    x=list(range(len(item_data))),
                    y=[mean_diff] * len(item_data),
                    name=f'{item} 평균 편차: {mean_diff:.3f}',
                    mode='lines',
                    line=dict(color=color, width=1.5, dash='dot')
                ))
                
                # 첫 번째 항목의 날짜 레이블만 사용
                if i == 0:
                    n_points = len(item_data)
                    n_labels = min(10, n_points)
                    
                    if n_points <= n_labels:
                        label_indices = list(range(n_points))
                    else:
                        step = n_points / n_labels
                        label_indices = [int(i * step) for i in range(n_labels)]
                        if label_indices[-1] != n_points - 1:
                            label_indices.append(n_points - 1)
                    
                    tickvals = label_indices
                    ticktext = [item_data.iloc[i]['날짜'].strftime('%y-%m-%d') for i in label_indices]
            
            # 0선 추가 (기준선)
            diff_fig.add_trace(go.Scatter(
                x=[0, len(item_data)-1],
                y=[0, 0],
                name='기준선',
                mode='lines',
                line=dict(color='black', width=1, dash='dash')
            ))
            
            # 레이아웃 설정
            diff_fig.update_layout(
                xaxis=dict(
                    title='측정 순서',
                    tickmode='array',
                    tickvals=tickvals,
                    ticktext=ticktext,
                    tickangle=45
                ),
                yaxis_title='편차 (실측-배합)',
                hovermode='x unified',
                showlegend=True,
                legend=dict(
                    orientation='h',
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                ),
                margin=dict(l=40, r=40, t=40, b=80),
                height=500
            )
        
        # 편차 그래프 표시
        st.plotly_chart(diff_fig, use_container_width=True)

        
    # 이상치 정보 및 부적합 정보 컨테이너 추가 - 그래프 아래에 위치
    anomaly_container = st.container()
    
    with anomaly_container:
        if len(composition_types) == 1:
            item_data = filtered_df[filtered_df['항목'] == composition_types[0]]
            item_name = composition_types[0]
            
            # 이상치 정보 표시
            st.markdown(f"""
            <div class="anomaly-box">
                <h4 style="margin-top: 0;">{item_name} 이상치 정보</h4>
            </div>
            """, unsafe_allow_html=True)
            
            # 이상치 테이블 생성
            mean = item_data['실측'].mean()
            std = item_data['실측'].std()
            outlier_points = item_data[abs(item_data['실측'] - mean) > sigma * std]
            
            if not outlier_points.empty:
                # 이상치 정보 테이블 생성
                outlier_table = outlier_points.copy()
                outlier_table = outlier_table.sort_values('날짜', ascending=False)
                outlier_table['날짜'] = outlier_table['날짜'].dt.strftime('%Y년 %m월 %d일')
                outlier_table['편차'] = outlier_table['실측'] - outlier_table['배합']
                
                # 필요한 열만 선택하고 이름 변경
                outlier_display = outlier_table[['날짜', '실측', '배합', '편차']].reset_index(drop=True)
                
                # 테이블 표시
                st.dataframe(outlier_display, use_container_width=True)
            else:
                st.info("설정된 관리한계(상한선/하한선)를 벗어나는 이상치가 없습니다.")
            
            # 부적합 정보 표시
            st.markdown(f"""
            <div class="incompatible-box">
                <h4 style="margin-top: 0;">{item_name} 부적합 정보</h4>
            </div>
            """, unsafe_allow_html=True)
            
            # 부적합 테이블 생성
            out_of_spec = item_data[
                (item_data['실측'] > item_data['상한선']) | 
                (item_data['실측'] < item_data['하한선'])
            ]
            
            if not out_of_spec.empty:
                # 부적합 정보 테이블 생성
                oos_table = out_of_spec.copy()
                oos_table = oos_table.sort_values('날짜', ascending=False)
                oos_table['날짜'] = oos_table['날짜'].dt.strftime('%Y년 %m월 %d일')
                oos_table['편차'] = oos_table['실측'] - oos_table['배합']
                
                # 필요한 열만 선택하고 이름 변경
                oos_display = oos_table[['날짜', '실측', '배합', '편차']].reset_index(drop=True)
                
                # 테이블 표시
                st.dataframe(oos_display, use_container_width=True)
            else:
                st.info("설정된 관리한계(상한선/하한선)를 벗어나는 부적합 항목이 없습니다.")
        
        else:
            # 다중 항목 선택 시 탭으로 구성
            tabs = st.tabs(composition_types)
            
            for i, (tab, item) in enumerate(zip(tabs, composition_types)):
                with tab:
                    item_data = filtered_df[filtered_df['항목'] == item]
                    
                    # 이상치 정보 표시
                    st.markdown(f"""
                    <div class="anomaly-box">
                        <h4 style="margin-top: 0;">{item} 이상치 정보</h4>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # 이상치 테이블 생성
                    mean = item_data['실측'].mean()
                    std = item_data['실측'].std()
                    outlier_points = item_data[abs(item_data['실측'] - mean) > sigma * std]
                    
                    if not outlier_points.empty:
                        # 이상치 정보 테이블 생성
                        outlier_table = outlier_points.copy()
                        outlier_table = outlier_table.sort_values('날짜', ascending=False)
                        outlier_table['날짜'] = outlier_table['날짜'].dt.strftime('%Y년 %m월 %d일')
                        outlier_table['편차'] = outlier_table['실측'] - outlier_table['배합']
                        
                        # 필요한 열만 선택하고 이름 변경
                        outlier_display = outlier_table[['날짜', '실측', '배합', '편차']].reset_index(drop=True)
                        
                        # 테이블 표시
                        st.dataframe(outlier_display, use_container_width=True)
                    else:
                        st.info("설정된 관리한계(상한선/하한선)를 벗어나는 이상치가 없습니다.")
                    
                    # 부적합 정보 표시
                    st.markdown(f"""
                    <div class="incompatible-box">
                        <h4 style="margin-top: 0;">{item} 부적합 정보</h4>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # 부적합 테이블 생성
                    out_of_spec = item_data[
                        (item_data['실측'] > item_data['상한선']) | 
                        (item_data['실측'] < item_data['하한선'])
                    ]
                    
                    if not out_of_spec.empty:
                        # 부적합 정보 테이블 생성
                        oos_table = out_of_spec.copy()
                        oos_table = oos_table.sort_values('날짜', ascending=False)
                        oos_table['날짜'] = oos_table['날짜'].dt.strftime('%Y년 %m월 %d일')
                        oos_table['편차'] = oos_table['실측'] - oos_table['배합']
                        
                        # 필요한 열만 선택하고 이름 변경
                        oos_display = oos_table[['날짜', '실측', '배합', '편차']].reset_index(drop=True)
                        
                        # 테이블 표시
                        st.dataframe(oos_display, use_container_width=True)
                    else:
                        st.info("설정된 관리한계(상한선/하한선)를 벗어나는 부적합 항목이 없습니다.")
    
    # 통계 컨테이너 - 이상치/부적합 정보 아래에 위치
    stats_container = st.container()
  
    with stats_container:
        if len(composition_types) == 1:
            # 단일 항목 선택 시 상세 통계
            st.markdown(f"""
            <div style="background-color: #f0f2f6; border-radius: 10px; padding: 15px; margin-top: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
                <h3 style="margin-top: 0;">상세 통계 정보</h3>
            </div>
            """, unsafe_allow_html=True)
          
            item_data = filtered_df[filtered_df['항목'] == composition_types[0]]
            item = composition_types[0]
          
            # 공정능력지수 계산
            if '상한선' in item_data.columns and '하한선' in item_data.columns:
                ucl = item_data['상한선'].mean()
                lcl = item_data['하한선'].mean()
              
                process_capability = calculate_process_capability(item_data['실측'], ucl, lcl, sigma)
              
                # 공정능력지수 표시
                st.markdown("#### 공정능력지수 (Process Capability)")
              
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric(label="Cp", value=f"{process_capability['Cp']:.3f}")
                with col2:
                    st.metric(label="Cpk", value=f"{process_capability['Cpk']:.3f}")
                with col3:
                    st.metric(label="Cpu", value=f"{process_capability['Cpu']:.3f}")
                with col4:
                    st.metric(label="Cpl", value=f"{process_capability['Cpl']:.3f}")
              
                # 예상 불량률 표시
                st.metric(label="예상 불량률 (PPM)", value=f"{process_capability['PPM']:.2f}")
          
            # 히스토그램 및 정규분포 그래프
            st.markdown(f"""
            <div style="background-color: #f8f9fa; border-radius: 10px; padding: 15px; margin-top: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
                <h3 style="margin-top: 0;">{item} 분포 분석</h3>
            </div>
            """, unsafe_allow_html=True)
          
            # 히스토그램과 정규분포 곡선
            hist_fig = go.Figure()
          
            # 히스토그램 생성
            hist_data = item_data['실측'].dropna()
            if len(hist_data) > 0:
                bins = min(20, max(5, int(np.sqrt(len(hist_data)))))  # 빈 개수 자동 계산
              
                # 히스토그램 트레이스 추가
                hist_fig.add_trace(go.Histogram(
                    x=hist_data,
                    name='실측 데이터',
                    opacity=0.7,
                    marker=dict(color='royalblue'),
                    nbinsx=bins,
                    histnorm='probability density'
                ))
              
                # 정규분포 곡선 추가
                mean = item_data['실측'].mean()
                std = item_data['실측'].std()
                x_range = np.linspace(hist_data.min() - 0.5, hist_data.max() + 0.5, 100)
                y_range = stats.norm.pdf(x_range, mean, std)
              
                hist_fig.add_trace(go.Scatter(
                    x=x_range,
                    y=y_range,
                    mode='lines',
                    name='정규분포',
                    line=dict(color='red', width=2)
                ))
              
                # 관리한계선 추가
                if '상한선' in item_data.columns:
                    ucl = item_data['상한선'].mean()
                    hist_fig.add_trace(go.Scatter(
                        x=[ucl, ucl],
                        y=[0, stats.norm.pdf(mean, mean, std) * 1.2],
                        mode='lines',
                        name='상한선',
                        line=dict(color='green', width=2, dash='dash')
                    ))
              
                if '하한선' in item_data.columns:
                    lcl = item_data['하한선'].mean()
                    hist_fig.add_trace(go.Scatter(
                        x=[lcl, lcl],
                        y=[0, stats.norm.pdf(mean, mean, std) * 1.2],
                        mode='lines',
                        name='하한선',
                        line=dict(color='green', width=2, dash='dash')
                    ))
              
                # 통계적 관리한계선 추가
                statistical_ucl = mean + sigma * std
                statistical_lcl = mean - sigma * std
              
                hist_fig.add_trace(go.Scatter(
                    x=[statistical_ucl, statistical_ucl],
                    y=[0, stats.norm.pdf(mean, mean, std) * 1.2],
                    mode='lines',
                    name=f'통계적 상한선 ({sigma}σ)',
                    line=dict(color='red', width=2, dash='dot')
                ))
              
                hist_fig.add_trace(go.Scatter(
                    x=[statistical_lcl, statistical_lcl],
                    y=[0, stats.norm.pdf(mean, mean, std) * 1.2],
                    mode='lines',
                    name=f'통계적 하한선 ({sigma}σ)',
                    line=dict(color='red', width=2, dash='dot')
                ))
              
                # 평균선 추가
                hist_fig.add_trace(go.Scatter(
                    x=[mean, mean],
                    y=[0, stats.norm.pdf(mean, mean, std) * 1.2],
                    mode='lines',
                    name='평균',
                    line=dict(color='black', width=2)
                ))
              
                # 그래프 레이아웃 설정
                hist_fig.update_layout(
                    title=f'{item} 분포 히스토그램',
                    xaxis_title='실측값',
                    yaxis_title='빈도 (확률 밀도)',
                    bargap=0.05,
                    bargroupgap=0.1,
                    height=400,
                    legend=dict(
                        yanchor="top",
                        y=0.99,
                        xanchor="right",
                        x=0.99
                    )
                )
              
                st.plotly_chart(hist_fig, use_container_width=True)

def plotly_chart_with_click_event(fig, key=None):
    """
    클릭 이벤트를 처리할 수 있는 Plotly 차트를 표시합니다.
    """
    # 클릭 이벤트 활성화
    fig.update_layout(clickmode='event+select')
    
    # 차트 표시
    chart = st.plotly_chart(fig, use_container_width=True, key=key)
    
    # 클릭 이벤트를 위한 컴포넌트 키 생성
    click_key = f"{key}_click" if key else "chart_click"
    
    # 세션 상태에 클릭 데이터 저장 공간 생성
    if click_key not in st.session_state:
        st.session_state[click_key] = []
    
    # Streamlit 컴포넌트 생성 (JavaScript 이벤트 처리용)
    # 주의: components.html()에는 key 매개변수를 전달하지 않음
    # components.html에서 key 매개변수 제거
    components.html(
        f"""
        <script>
            const chart = document.querySelector('#{key} .js-plotly-plot');
            if (chart) {{
                chart.on('plotly_click', function(data) {{
                    // ... (스크립트 내용)
                }});
            }}
        </script>
        """,
        height=0,
        width=0
    )
    
    # 세션 상태에서 클릭 데이터 반환
    return st.session_state.get(click_key, [])



def display_overview(all_data, filtered_df, selected_sheet, sigma):
    """
    전체 현황을 표시하는 함수
    """
    st.subheader(f"{selected_sheet} 조성 전체 현황")
    
    # 전체 항목 가져오기
    all_items = filtered_df['항목'].unique()
    
    # 배합-실측 차이 분석 추가
    st.markdown(f"""
    <div style="background-color: #f0f2f6; border-radius: 10px; padding: 15px; margin-top: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
        <h3 style="margin-top: 0;">배합-실측 차이 분석</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # 각 항목별 배합-실측 차이 계산
    diff_stats = []
    
    for item in all_items:
        item_data = filtered_df[filtered_df['항목'] == item]
        
        # 편차 계산
        item_data['편차'] = item_data['실측'] - item_data['배합']
        
        # 통계치 계산
        mean_diff = item_data['편차'].mean()
        std_diff = item_data['편차'].std()
        
        # t-검정으로 통계적 유의성 확인
        t_stat, p_value = stats.ttest_1samp(item_data['편차'], 0)
        
        # 통계적 판단 (p < 0.05이면 통계적으로 유의한 차이)
        if p_value < 0.05:
            if mean_diff > 0:
                statistical_judgment = "실측값이 배합값보다 통계적으로 유의하게 높음"
                color = "red"
                technical_judgment = "배합 설정값이 실제 투입량보다 낮게 설정되어 있거나, 측정 과정에서 양의 편향 가능성"
            else:
                statistical_judgment = "실측값이 배합값보다 통계적으로 유의하게 낮음"
                color = "blue"
                technical_judgment = "공정 중 원료 손실 가능성 또는 원료 투입량이 설정값보다 적을 가능성"
        else:
            statistical_judgment = "실측값과 배합값 사이에 통계적으로 유의한 차이 없음"
            color = "green"
            technical_judgment = "배합과 실측이 잘 일치함"
        
        # 결과 저장
        diff_stats.append({
            '항목': item,
            '평균_편차': mean_diff,
            '편차_표준편차': std_diff,
            'p값': p_value,
            '통계적_판단': statistical_judgment,
            '기술적_판단': technical_judgment,
            '색상': color
        })
    
    # 데이터프레임으로 변환
    diff_stats_df = pd.DataFrame(diff_stats)
    
    # 통계적으로 유의한 차이가 있는 항목 필터링
    higher_items = diff_stats_df[(diff_stats_df['p값'] < 0.05) & (diff_stats_df['평균_편차'] > 0)]
    lower_items = diff_stats_df[(diff_stats_df['p값'] < 0.05) & (diff_stats_df['평균_편차'] < 0)]
    
    # 기술적 제언 표시 부분을 수정
    st.markdown("<h3 style='margin-bottom: 0.5rem;'>기술적 제언</h3>", unsafe_allow_html=True)

    # CSS 스타일 추가 - 카드 스타일과 높이 통일, 간격 조절
    st.markdown("""
    <style>
    .suggestion-container {
        display: flex;
        flex-direction: column;
        height: 100%;
        margin-top: 0.5rem; /* 상단 여백 줄임 */
    }
    .suggestion-card {
        background-color: #f8f9fa;
        border-radius: 10px;
        padding: 15px;
        margin-top: 0; /* 카드 상단 여백 제거 */
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        height: 100%;
        flex-grow: 1;
    }
    .stColumns {
        height: auto !important;
        margin-top: 0.5rem !important; /* 컬럼 상단 여백 줄임 */
    }
    .stColumn > div {
        height: 100%;
    }
    /* 추가: 헤딩과 컨텐츠 사이 간격 조절 */
    h3 {
        margin-bottom: 0.5rem !important;
    }
    /* 추가: 기본 마진 오버라이드 */
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 1rem !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Streamlit의 columns 기능 사용 - 간격 조절
    col1, col2 = st.columns(2, gap="small")  # gap을 small로 변경

    # 왼쪽 열에 실측값이 배합값보다 높은 항목 표시
    with col1:
        st.markdown('<div class="suggestion-container">', unsafe_allow_html=True)
        if len(higher_items) > 0:
            st.markdown(f"""
            <div class="suggestion-card">
                <h4 style="margin-top: 0; margin-bottom: 0.5rem;">실측값이 배합값보다 높은 항목 ({len(higher_items)}개)</h4>
                <p style="margin-top: 0.5rem;">다음 항목들은 실측값이 배합값보다 통계적으로 유의하게 높게 나타납니다:</p>
                <ul style="margin-top: 0.5rem; margin-bottom: 0.5rem;">
                    {"".join([f"<li><strong>{row['항목']}</strong>: 평균 편차 {row['평균_편차']:.3f}</li>" for _, row in higher_items.iterrows()])}
                </ul>
                <p style="margin-top: 0.5rem; margin-bottom: 0.5rem;"><strong>가능한 원인:</strong></p>
                <ul style="margin-top: 0.5rem; margin-bottom: 0.5rem;">
                    <li>배합 설정값이 실제 투입량보다 낮게 설정되어 있을 가능성</li>
                    <li>측정 과정에서의 양의 편향 가능성</li>
                    <li>원료 투입 과정에서의 초과 투입 가능성</li>
                </ul>
                <p style="margin-top: 0.5rem;"><strong>제언:</strong> 해당 항목들의 배합 설정값을 검토하고, 필요시 조정을 고려하세요.</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="suggestion-card">
                <h4 style="margin-top: 0; margin-bottom: 0.5rem;">실측값이 배합값보다 높은 항목</h4>
                <p style="margin-top: 0.5rem;">실측값이 배합값보다 통계적으로 유의하게 높은 항목이 없습니다.</p>
            </div>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # 오른쪽 열에 실측값이 배합값보다 낮은 항목 표시
    with col2:
        st.markdown('<div class="suggestion-container">', unsafe_allow_html=True)
        if len(lower_items) > 0:
            st.markdown(f"""
            <div class="suggestion-card">
                <h4 style="margin-top: 0; margin-bottom: 0.5rem;">실측값이 배합값보다 낮은 항목 ({len(lower_items)}개)</h4>
                <p style="margin-top: 0.5rem;">다음 항목들은 실측값이 배합값보다 통계적으로 유의하게 낮게 나타납니다:</p>
                <ul style="margin-top: 0.5rem; margin-bottom: 0.5rem;">
                    {"".join([f"<li><strong>{row['항목']}</strong>: 평균 편차 {row['평균_편차']:.3f}</li>" for _, row in lower_items.iterrows()])}
                </ul>
                <p style="margin-top: 0.5rem; margin-bottom: 0.5rem;"><strong>가능한 원인:</strong></p>
                <ul style="margin-top: 0.5rem; margin-bottom: 0.5rem;">
                    <li>공정 중 원료 손실 가능성</li>
                    <li>측정 과정에서의 음의 편향 가능성</li>
                    <li>원료 투입량이 설정값보다 적을 가능성</li>
                </ul>
                <p style="margin-top: 0.5rem;"><strong>제언:</strong> 해당 항목들의 원료 투입 과정과 측정 방법을 검토하세요.</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="suggestion-card">
                <h4 style="margin-top: 0; margin-bottom: 0.5rem;">실측값이 배합값보다 낮은 항목</h4>
                <p style="margin-top: 0.5rem;">실측값이 배합값보다 통계적으로 유의하게 낮은 항목이 없습니다.</p>
            </div>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)





    
    # 이상치 및 부적합 요약 정보 (기존 코드)
    st.markdown(f"""
    <div style="background-color: #f0f2f6; border-radius: 10px; padding: 15px; margin-top: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
        <h3 style="margin-top: 0;">이상치 및 부적합 요약</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # 모든 항목의 이상치 및 부적합 계산 (이하 기존 코드)
    anomaly_data = []

    
    for item in all_items:
        item_data = filtered_df[filtered_df['항목'] == item]
        
        # 통계치 계산
        mean = item_data['실측'].mean()
        std = item_data['실측'].std()
        
        # 이상치 계산
        outliers = item_data[abs(item_data['실측'] - mean) > sigma * std]
        
        # 부적합 계산
        out_of_spec = item_data[
            (item_data['실측'] > item_data['상한선']) | 
            (item_data['실측'] < item_data['하한선'])
        ]
       # 이상치 데이터 저장
        for _, row in outliers.iterrows():
            anomaly_data.append({
                '항목': item,
                '날짜': row['날짜'],
                '실측값': row['실측'],
                '배합값': row['배합'],
                '편차': row['실측'] - row['배합'],
                '평균': mean,
                '표준편차': std,
                '상한선': row['상한선'],
                '하한선': row['하한선'],
                '유형': '이상치',
                '비고': f"{abs(row['실측'] - mean) / std:.2f}σ 이탈"
            })
        
        # 부적합 데이터 저장 (이상치가 아닌 경우만)
        for _, row in out_of_spec.iterrows():
            if abs(row['실측'] - mean) <= sigma * std:  # 이상치가 아닌 경우만 추가
                anomaly_data.append({
                    '항목': item,
                    '날짜': row['날짜'],
                    '실측값': row['실측'],
                    '배합값': row['배합'],
                    '편차': row['실측'] - row['배합'],
                    '평균': mean,
                    '표준편차': std,
                    '상한선': row['상한선'],
                    '하한선': row['하한선'],
                    '유형': '부적합',
                    '비고': '규격 이탈'
                })
    
    # 데이터프레임으로 변환
    if anomaly_data:
        anomaly_df = pd.DataFrame(anomaly_data)
        anomaly_df['날짜'] = pd.to_datetime(anomaly_df['날짜'])
        anomaly_df = anomaly_df.sort_values('날짜', ascending=False)
        
        # 요약 통계 표시
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric(label="총 이상치 수", value=len(anomaly_df[anomaly_df['유형'] == '이상치']))
        with col2:
            st.metric(label="총 부적합 수", value=len(anomaly_df[anomaly_df['유형'] == '부적합']))
        with col3:
            st.metric(label="총 항목 수", value=len(all_items))
        
        # 이상치 및 부적합 항목별 분포 차트
        st.subheader("항목별 이상치 및 부적합 분포")
        
        # 항목별 이상치 및 부적합 개수 계산
        item_counts = anomaly_df.groupby(['항목', '유형']).size().reset_index(name='개수')
        
        # 차트 생성
        fig = px.bar(
            item_counts, 
            x='항목', 
            y='개수', 
            color='유형',
            barmode='group',
            color_discrete_map={'이상치': 'red', '부적합': 'blue'},
            title="항목별 이상치 및 부적합 개수"
        )
        
        fig.update_layout(
            xaxis_title="항목",
            yaxis_title="개수",
            legend_title="유형",
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # 이상치 및 부적합 시간 추이 차트
        st.subheader("이상치 및 부적합 시간 추이")
        
        # 날짜별 이상치 및 부적합 개수 계산
        anomaly_df['날짜_일자'] = anomaly_df['날짜'].dt.date
        date_counts = anomaly_df.groupby(['날짜_일자', '유형']).size().reset_index(name='개수')
        
        # 차트 생성
        fig = px.line(
            date_counts, 
            x='날짜_일자', 
            y='개수', 
            color='유형',
            markers=True,
            color_discrete_map={'이상치': 'red', '부적합': 'blue'},
            title="날짜별 이상치 및 부적합 발생 추이"
        )
        
        fig.update_layout(
            xaxis_title="날짜",
            yaxis_title="개수",
            legend_title="유형",
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # 이상치 및 부적합 데이터 테이블
        st.subheader("이상치 및 부적합 데이터 목록")
        
        # 표시할 컬럼 선택
        display_cols = ['항목', '날짜', '실측값', '배합값', '편차', '상한선', '하한선', '유형', '비고']
        
        # 날짜 포맷 변경
        anomaly_df['날짜'] = anomaly_df['날짜'].dt.strftime('%Y-%m-%d')
        
        # 데이터 테이블 표시
        st.dataframe(
            anomaly_df[display_cols], 
            use_container_width=True,
            hide_index=True,
            column_config={
                '날짜': st.column_config.DateColumn('날짜'),
                '실측값': st.column_config.NumberColumn('실측값', format="%.3f"),
                '배합값': st.column_config.NumberColumn('배합값', format="%.3f"),
                '편차': st.column_config.NumberColumn('편차', format="%.3f"),
                '상한선': st.column_config.NumberColumn('상한선', format="%.3f"),
                '하한선': st.column_config.NumberColumn('하한선', format="%.3f")
            }
        )
        
        # 데이터 행 선택 기능
        st.markdown("### 상세 분석")
        st.write("아래 데이터 행을 선택하여 상세 분석을 확인할 수 있습니다.")
        
        # 선택 가능한 데이터 목록 생성
        selection_data = anomaly_df[['항목', '날짜', '실측값', '유형']].copy()
        selection_data['표시'] = selection_data.apply(
            lambda row: f"{row['항목']} - {row['날짜']} ({row['유형']})", axis=1
        )
        
        # 선택 위젯
        selected_row = st.selectbox(
            "분석할 데이터 선택",
            options=selection_data['표시'].tolist(),
            index=0 if not selection_data.empty else None
        )
        
        if selected_row and not selection_data.empty:  # 조건 추가
            # 선택된 행 찾기
            selected_idx = selection_data[selection_data['표시'] == selected_row].index[0]
            selected_item = selection_data.loc[selected_idx, '항목']
            selected_date = selection_data.loc[selected_idx, '날짜']
            
            if st.button(f"'{selected_item}' 상세 분석 보기"):
                st.session_state.selected_item = selected_item
                st.session_state.tab_selection = "상세 분석"
                st.rerun()
                
    else:
        st.info("선택한 기간 내에 이상치 또는 부적합 데이터가 없습니다.")


        
def register_click_callback():
    """
    클릭 이벤트 콜백을 등록하는 함수
    """
    components.html(
        """
        <script>
            // Plotly 차트에 클릭 이벤트 리스너 추가
            const observer = new MutationObserver(function(mutations) {
                mutations.forEach(function(mutation) {
                    if (mutation.addedNodes.length) {
                        const charts = document.querySelectorAll('.js-plotly-plot');
                        charts.forEach(chart => {
                            if (!chart.hasAttribute('data-click-registered')) {
                                chart.setAttribute('data-click-registered', 'true');
                                chart.on('plotly_click', function(data) {
                                    const points = data.points[0];
                                    const clickData = {
                                        curveNumber: points.curveNumber,
                                        pointIndex: points.pointIndex,
                                        x: points.x,
                                        y: points.y
                                    };
                                    
                                    // Streamlit에 이벤트 전달
                                    window.parent.postMessage({
                                        type: 'streamlit:setComponentValue',
                                        value: JSON.stringify(clickData),
                                        dataType: 'json'
                                    }, '*');
                                });
                            }
                        });
                    }
                });
            });
            
            // DOM 변화 감시 시작
            observer.observe(document.body, { childList: true, subtree: true });
        </script>
        """,
        height=0,
        width=0
    )


if __name__ == "__main__":
    main()